
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook

# 要遍历的根目录
ROOT_DIR = "/data/home/aaayeyeye/config/GameData/xls"
TARGET_TEXT = "典藏球星自选礼盒"

def scan_xlsx_files(root_dir):
    print(root_dir)
    for root, dirs, files in os.walk(root_dir):
        print(dirs)
        for filename in files:
            if filename.endswith(".xlsx"):
                full_path = os.path.join(root, filename)
                scan_workbook(full_path)

def scan_workbook(path):
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"无法读取文件 {path}: {e}")
        return
    print(path)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        if sheet_contains_text(sheet, TARGET_TEXT):
            print(f"发现匹配：文件《{os.path.basename(path)}》 → sheet《{sheetname}》")

def sheet_contains_text(sheet, target):
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and target in cell:
                return True
    return False


if __name__ == "__main__":
    scan_xlsx_files(ROOT_DIR)

