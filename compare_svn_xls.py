#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SVN XLS 文件对比工具
支持双列diff对比、版本切换、Sheet切换、彩色高亮显示
"""

import os
import sys
import tempfile
import subprocess
import re
from typing import List, Tuple, Optional, Dict
import curses
from dataclasses import dataclass

try:
    import xlrd
    import openpyxl
except ImportError:
    print("请先安装依赖: pip install xlrd openpyxl")
    sys.exit(1)


@dataclass
class CellData:
    """单元格数据"""
    value: str
    changed: bool = False
    
    def __str__(self):
        return str(self.value) if self.value is not None else ""


class SVNHelper:
    """SVN 操作辅助类"""
    
    @staticmethod
    def get_file_history(file_path: str, limit: int = 10) -> List[Tuple[int, str]]:
        """获取文件的版本历史"""
        try:
            cmd = ['svn', 'log', '-q', '-l', str(limit), file_path]
            result = subprocess.run(cmd, capture_output=True, text=True, check=True)
            
            revisions = []
            for line in result.stdout.split('\n'):
                match = re.match(r'^r(\d+)\s+\|\s+(\S+)', line)
                if match:
                    rev_num = int(match.group(1))
                    author = match.group(2)
                    revisions.append((rev_num, author))
            
            return revisions
        except subprocess.CalledProcessError as e:
            print(f"获取SVN历史失败: {e}")
            return []
    
    @staticmethod
    def export_file_at_revision(file_path: str, revision: int) -> Optional[str]:
        """导出指定版本的文件到临时目录"""
        try:
            temp_dir = tempfile.gettempdir()
            file_name = os.path.basename(file_path)
            temp_file = os.path.join(temp_dir, f"{file_name}.r{revision}")
            
            cmd = ['svn', 'export', '-r', str(revision), file_path, temp_file]
            subprocess.run(cmd, capture_output=True, check=True)
            
            return temp_file
        except subprocess.CalledProcessError as e:
            print(f"导出文件失败: {e}")
            return None


class XLSReader:
    """XLS/XLSX 文件读取器"""
    
    @staticmethod
    def read_file(file_path: str) -> Dict[str, List[List[str]]]:
        """读取 XLS/XLSX 文件，返回所有 sheet 的数据"""
        sheets_data = {}
        
        # 尝试使用 openpyxl 读取 xlsx
        if file_path.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append([str(cell) if cell is not None else "" for cell in row])
                    sheets_data[sheet_name] = data
                wb.close()
                return sheets_data
            except Exception as e:
                print(f"使用 openpyxl 读取失败: {e}")
        
        # 尝试使用 xlrd 读取 xls
        try:
            wb = xlrd.open_workbook(file_path)
            for sheet in wb.sheets():
                data = []
                for row_idx in range(sheet.nrows):
                    row = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        row.append(str(cell.value) if cell.value != '' else "")
                    data.append(row)
                sheets_data[sheet.name] = data
            return sheets_data
        except Exception as e:
            print(f"读取文件失败: {e}")
            return {}


class DiffViewer:
    """双列 Diff 查看器"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.revisions: List[Tuple[int, str]] = []
        self.current_rev_idx = 0  # 当前对比的版本索引对 (0表示最新vs上一个)
        
        self.left_data: Dict[str, List[List[CellData]]] = {}
        self.right_data: Dict[str, List[List[CellData]]] = {}
        self.sheet_names: List[str] = []
        self.current_sheet_idx = 0
        
        self.offset_x = 0  # 水平滚动偏移
        self.offset_y = 0  # 垂直滚动偏移
        self.col_width = 20  # 列宽
        
        # 当前 diff 位置缓存
        self.diff_positions: List[Tuple[int, int]] = []  # (row, col) 列表
        self.current_diff_idx = -1  # 当前 diff 索引
        
        # 颜色对
        self.COLOR_CHANGED = 1
        self.COLOR_HEADER = 2
        self.COLOR_NORMAL = 3
        self.COLOR_STATUS = 4
    
    def init_colors(self):
        """初始化颜色方案"""
        curses.init_pair(self.COLOR_CHANGED, curses.COLOR_BLACK, curses.COLOR_YELLOW)
        curses.init_pair(self.COLOR_HEADER, curses.COLOR_WHITE, curses.COLOR_BLUE)
        curses.init_pair(self.COLOR_NORMAL, curses.COLOR_WHITE, curses.COLOR_BLACK)
        curses.init_pair(self.COLOR_STATUS, curses.COLOR_BLACK, curses.COLOR_CYAN)
    
    def load_revisions(self) -> bool:
        """加载版本历史"""
        self.revisions = SVNHelper.get_file_history(self.file_path, limit=20)
        if len(self.revisions) < 2:
            print("文件版本历史不足，需要至少2个版本")
            return False
        return True
    
    def load_data(self, rev_a_idx: int, rev_b_idx: int) -> bool:
        """加载两个版本的数据进行对比"""
        if rev_a_idx >= len(self.revisions) or rev_b_idx >= len(self.revisions):
            return False
        
        rev_a = self.revisions[rev_a_idx][0]
        rev_b = self.revisions[rev_b_idx][0]
        
        # 导出文件
        file_a = SVNHelper.export_file_at_revision(self.file_path, rev_a)
        file_b = SVNHelper.export_file_at_revision(self.file_path, rev_b)
        
        if not file_a or not file_b:
            return False
        
        # 读取数据
        sheets_a = XLSReader.read_file(file_a)
        sheets_b = XLSReader.read_file(file_b)
        
        # 清理临时文件
        try:
            os.remove(file_a)
            os.remove(file_b)
        except:
            pass
        
        if not sheets_a or not sheets_b:
            return False
        
        # 获取所有 sheet 名称
        self.sheet_names = list(set(sheets_a.keys()) | set(sheets_b.keys()))
        self.sheet_names.sort()
        
        # 对比数据并标记变更
        self.left_data = {}
        self.right_data = {}
        
        for sheet_name in self.sheet_names:
            data_a = sheets_a.get(sheet_name, [])
            data_b = sheets_b.get(sheet_name, [])
            
            max_rows = max(len(data_a), len(data_b))
            max_cols = max(
                max([len(row) for row in data_a] + [0]),
                max([len(row) for row in data_b] + [0])
            )
            
            left_sheet = []
            right_sheet = []
            
            for row_idx in range(max_rows):
                left_row = []
                right_row = []
                
                for col_idx in range(max_cols):
                    val_a = data_a[row_idx][col_idx] if row_idx < len(data_a) and col_idx < len(data_a[row_idx]) else ""
                    val_b = data_b[row_idx][col_idx] if row_idx < len(data_b) and col_idx < len(data_b[row_idx]) else ""
                    
                    changed = (val_a != val_b)
                    
                    left_row.append(CellData(val_a, changed))
                    right_row.append(CellData(val_b, changed))
                
                left_sheet.append(left_row)
                right_sheet.append(right_row)
            
            self.left_data[sheet_name] = left_sheet
            self.right_data[sheet_name] = right_sheet
        
        # 更新 diff 位置列表
        self.update_diff_positions()
        
        return True
    
    def update_diff_positions(self):
        """更新当前 Sheet 的所有 diff 位置"""
        self.diff_positions = []
        
        if not self.sheet_names:
            return
        
        current_sheet = self.sheet_names[self.current_sheet_idx]
        left_sheet = self.left_data.get(current_sheet, [])
        
        for row_idx, row in enumerate(left_sheet):
            for col_idx, cell in enumerate(row):
                if cell.changed:
                    self.diff_positions.append((row_idx, col_idx))
        
        self.current_diff_idx = -1  # 重置当前 diff 索引
    
    def goto_next_diff(self):
        """跳转到下一个 diff 位置"""
        if not self.diff_positions:
            return
        
        # 如果还没有选中任何 diff，从当前视图位置后第一个开始
        if self.current_diff_idx == -1:
            # 查找当前视图位置之后的第一个 diff
            for idx, (row, col) in enumerate(self.diff_positions):
                if row > self.offset_y or (row == self.offset_y and col >= self.offset_x):
                    self.current_diff_idx = idx
                    break
            
            # 如果没找到，则从第一个开始
            if self.current_diff_idx == -1:
                self.current_diff_idx = 0
        else:
            # 移动到下一个 diff
            self.current_diff_idx = (self.current_diff_idx + 1) % len(self.diff_positions)
        
        # 更新视图位置
        row, col = self.diff_positions[self.current_diff_idx]
        self.offset_y = row
        self.offset_x = col
    
    def goto_prev_diff(self):
        """跳转到上一个 diff 位置"""
        if not self.diff_positions:
            return
        
        # 如果还没有选中任何 diff，从当前视图位置前一个开始
        if self.current_diff_idx == -1:
            # 查找当前视图位置之前的最后一个 diff
            for idx in range(len(self.diff_positions) - 1, -1, -1):
                row, col = self.diff_positions[idx]
                if row < self.offset_y or (row == self.offset_y and col < self.offset_x):
                    self.current_diff_idx = idx
                    break
            
            # 如果没找到，则从最后一个开始
            if self.current_diff_idx == -1:
                self.current_diff_idx = len(self.diff_positions) - 1
        else:
            # 移动到上一个 diff
            self.current_diff_idx = (self.current_diff_idx - 1) % len(self.diff_positions)
        
        # 更新视图位置
        row, col = self.diff_positions[self.current_diff_idx]
        self.offset_y = row
        self.offset_x = col
    
    def draw_screen(self, stdscr):
        """绘制屏幕"""
        stdscr.clear()
        height, width = stdscr.getmaxyx()
        
        # 确保终端窗口足够大
        if height < 10 or width < 40:
            try:
                stdscr.addstr(0, 0, "终端窗口太小，请调整大小")
            except:
                pass
            stdscr.refresh()
            return
        
        if not self.sheet_names:
            stdscr.addstr(0, 0, "没有数据可显示")
            stdscr.refresh()
            return
        
        current_sheet = self.sheet_names[self.current_sheet_idx]
        left_sheet = self.left_data[current_sheet]
        right_sheet = self.right_data[current_sheet]
        
        rev_a = self.revisions[self.current_rev_idx][0]
        rev_b = self.revisions[self.current_rev_idx + 1][0]
        
        # 绘制标题栏 (行0)
        title = f"Sheet: {current_sheet} ({self.current_sheet_idx + 1}/{len(self.sheet_names)})"
        version_info = f"r{rev_a} vs r{rev_b}"
        stdscr.attron(curses.color_pair(self.COLOR_HEADER))
        try:
            left_part = title[:width // 2 - 1].ljust(width // 2 - 1)
            stdscr.addstr(0, 0, left_part)
        except curses.error:
            pass
        try:
            right_part = version_info[:width // 2].ljust(width // 2)
            stdscr.addstr(0, width // 2, right_part)
        except curses.error:
            pass
        stdscr.attroff(curses.color_pair(self.COLOR_HEADER))
        
        # 绘制列标题 (行1)
        half_width = width // 2
        visible_cols = half_width // self.col_width
        
        stdscr.attron(curses.color_pair(self.COLOR_STATUS))
        left_header = f"  r{rev_a} (新版本)"
        right_header = f"  r{rev_b} (旧版本)"
        try:
            stdscr.addstr(1, 0, left_header[:half_width - 1].ljust(half_width - 1))
        except curses.error:
            pass
        try:
            stdscr.addstr(1, half_width, right_header[:half_width].ljust(half_width))
        except curses.error:
            pass
        stdscr.attroff(curses.color_pair(self.COLOR_STATUS))
        
        # 绘制数据区域 (从行2开始)
        data_start_row = 2
        visible_rows = height - data_start_row - 1  # 留一行给状态栏
        
        for screen_row in range(visible_rows):
            data_row = self.offset_y + screen_row
            
            if data_row >= len(left_sheet):
                break
            
            # 绘制左侧 (新版本)
            for screen_col in range(visible_cols):
                data_col = self.offset_x + screen_col
                
                if data_col >= len(left_sheet[data_row]):
                    break
                
                cell = left_sheet[data_row][data_col]
                x_pos = screen_col * self.col_width
                y_pos = data_start_row + screen_row
                
                content = str(cell)[:self.col_width - 1].ljust(self.col_width - 1)
                
                try:
                    if cell.changed:
                        stdscr.attron(curses.color_pair(self.COLOR_CHANGED))
                        stdscr.addstr(y_pos, x_pos, content)
                        stdscr.attroff(curses.color_pair(self.COLOR_CHANGED))
                    else:
                        stdscr.addstr(y_pos, x_pos, content)
                except curses.error:
                    pass
            
            # 绘制分隔线
            try:
                stdscr.addstr(data_start_row + screen_row, half_width - 1, "|")
            except curses.error:
                pass
            
            # 绘制右侧 (旧版本)
            for screen_col in range(visible_cols):
                data_col = self.offset_x + screen_col
                
                if data_col >= len(right_sheet[data_row]):
                    break
                
                cell = right_sheet[data_row][data_col]
                x_pos = half_width + screen_col * self.col_width
                y_pos = data_start_row + screen_row
                
                content = str(cell)[:self.col_width - 1].ljust(self.col_width - 1)
                
                try:
                    if cell.changed:
                        stdscr.attron(curses.color_pair(self.COLOR_CHANGED))
                        stdscr.addstr(y_pos, x_pos, content)
                        stdscr.attroff(curses.color_pair(self.COLOR_CHANGED))
                    else:
                        stdscr.addstr(y_pos, x_pos, content)
                except curses.error:
                    pass
        
        # 绘制状态栏 (最后一行)
        status_row = height - 1
        
        # 计算 diff 统计信息
        diff_info = f"Diff: {len(self.diff_positions)}"
        if self.diff_positions and self.current_diff_idx >= 0:
            diff_info += f" (当前: {self.current_diff_idx + 1}/{len(self.diff_positions)})"
        
        status = f"位置: 行{self.offset_y + 1}, 列{self.offset_x + 1} | {diff_info} | "
        status += f"方向键:移动 Tab:切换Sheet ]/[:跳转Diff N/P:切换版本 V:输入版本 Q:退出"
        
        stdscr.attron(curses.color_pair(self.COLOR_STATUS))
        try:
            # 确保状态栏文字不超过宽度限制，并且不会在最后一个字符位置写入
            status_text = status[:width - 1].ljust(width - 1)
            stdscr.addstr(status_row, 0, status_text)
        except curses.error:
            pass
        stdscr.attroff(curses.color_pair(self.COLOR_STATUS))
        
        stdscr.refresh()
    
    def get_version_input(self, stdscr) -> Optional[Tuple[int, int]]:
        """获取用户输入的版本号"""
        height, width = stdscr.getmaxyx()
        
        # 显示版本列表
        curses.echo()
        stdscr.clear()
        stdscr.addstr(0, 0, "可用版本列表:")
        
        for idx, (rev, author) in enumerate(self.revisions[:10]):
            stdscr.addstr(idx + 2, 2, f"{idx}: r{rev} by {author}")
        
        stdscr.addstr(13, 0, "请输入版本A的索引 (0为最新): ")
        stdscr.refresh()
        
        try:
            idx_a_str = stdscr.getstr(13, 32, 3).decode('utf-8')
            idx_a = int(idx_a_str)
            
            stdscr.addstr(14, 0, "请输入版本B的索引 (应大于版本A): ")
            stdscr.refresh()
            
            idx_b_str = stdscr.getstr(14, 36, 3).decode('utf-8')
            idx_b = int(idx_b_str)
            
            if 0 <= idx_a < len(self.revisions) and 0 <= idx_b < len(self.revisions) and idx_a < idx_b:
                curses.noecho()
                return (idx_a, idx_b)
        except:
            pass
        
        curses.noecho()
        return None
    
    def run(self, stdscr):
        """运行主循环"""
        curses.curs_set(0)  # 隐藏光标
        curses.start_color()
        curses.use_default_colors()
        self.init_colors()
        
        # 加载初始数据
        if not self.load_revisions():
            return
        
        if not self.load_data(0, 1):
            stdscr.addstr(0, 0, "加载数据失败，按任意键退出")
            stdscr.getch()
            return
        
        while True:
            self.draw_screen(stdscr)
            
            key = stdscr.getch()
            
            if key == ord('q') or key == ord('Q'):
                break
            
            elif key == curses.KEY_UP:
                self.offset_y = max(0, self.offset_y - 1)
            
            elif key == curses.KEY_DOWN:
                current_sheet = self.sheet_names[self.current_sheet_idx]
                max_rows = len(self.left_data[current_sheet])
                self.offset_y = min(max_rows - 1, self.offset_y + 1)
            
            elif key == curses.KEY_LEFT:
                self.offset_x = max(0, self.offset_x - 1)
            
            elif key == curses.KEY_RIGHT:
                current_sheet = self.sheet_names[self.current_sheet_idx]
                max_cols = len(self.left_data[current_sheet][0]) if self.left_data[current_sheet] else 0
                self.offset_x = min(max_cols - 1, self.offset_x + 1)
            
            elif key == ord('\t') or key == ord('s') or key == ord('S'):
                # Tab 或 S 键切换 Sheet
                self.current_sheet_idx = (self.current_sheet_idx + 1) % len(self.sheet_names)
                self.offset_x = 0
                self.offset_y = 0
                self.update_diff_positions()  # 切换 Sheet 后更新 diff 位置
            
            elif key == ord(']'):
                # ] 键跳转到下一个 diff 位置
                self.goto_next_diff()
            
            elif key == ord('['):
                # [ 键跳转到上一个 diff 位置
                self.goto_prev_diff()
            
            elif key == ord('n') or key == ord('N'):
                # N 键切换到下一个版本对比
                if self.current_rev_idx + 2 < len(self.revisions):
                    self.current_rev_idx += 1
                    if self.load_data(self.current_rev_idx, self.current_rev_idx + 1):
                        self.offset_x = 0
                        self.offset_y = 0
                        self.current_sheet_idx = 0
                        self.update_diff_positions()  # 切换版本后更新 diff 位置
            
            elif key == ord('p') or key == ord('P'):
                # P 键切换到上一个版本对比
                if self.current_rev_idx > 0:
                    self.current_rev_idx -= 1
                    if self.load_data(self.current_rev_idx, self.current_rev_idx + 1):
                        self.offset_x = 0
                        self.offset_y = 0
                        self.current_sheet_idx = 0
                        self.update_diff_positions()  # 切换版本后更新 diff 位置
            
            elif key == ord('v') or key == ord('V'):
                # V 键输入版本号
                result = self.get_version_input(stdscr)
                if result:
                    idx_a, idx_b = result
                    if self.load_data(idx_a, idx_b):
                        self.current_rev_idx = idx_a
                        self.offset_x = 0
                        self.offset_y = 0
                        self.current_sheet_idx = 0
                        self.update_diff_positions()  # 切换版本后更新 diff 位置


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("用法: python svn_xls_diff.py <xls文件路径>")
        print("\n快捷键说明:")
        print("  方向键: 上下左右移动")
        print("  Tab/S: 切换 Sheet")
        print("  ]: 跳转到下一个 Diff 位置")
        print("  [: 跳转到上一个 Diff 位置")
        print("  N: 切换到下一个版本对比 (例如: r2 vs r1)")
        print("  P: 切换到上一个版本对比 (例如: 回到 r3 vs r2)")
        print("  V: 输入自定义版本号进行对比")
        print("  Q: 退出")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        sys.exit(1)
    
    viewer = DiffViewer(file_path)
    
    try:
        curses.wrapper(viewer.run)
    except KeyboardInterrupt:
        print("\n已取消")
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
